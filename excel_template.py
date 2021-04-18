from openpyxl import Workbook

from sequence_stats import get_distribution_histograms, base_distances, pair_distances, get_base_sequences_parts
from statistics import mean, mode

import numpy as np

from wild_type import get_wild_type

max_distance = 20


def get_sequences_by_region(database, region):
    sequences_ref = database.collection(u'sequences')
    sequences_objs = sequences_ref.where(u'seq_len', u'<=', 377).where(u'regions', u'array_contains_any',
                                                                       [region]).stream()
    sequences = [obj.to_dict()['seq'] for obj in sequences_objs]
    return sequences


def filter_base_sequences(var):
    if var.to_dict()['seq_len'] <= 377:
        return True
    return False


def get_sequences_not_in_regions(database, region):
    sequences_ref = database.collection(u'sequences')
    sequences_objs = sequences_ref.where(u'regions', u'not-in', [region]).stream()
    sequences_objs = filter(filter_base_sequences, sequences_objs)
    sequences = [obj.to_dict()['seq'] for obj in sequences_objs]
    return sequences


def get_wild_type_by_region(database, region):
    wild_type_ref = database.collection(u'wild_type')
    wild_type = wild_type_ref.where(u'regions', u'array_contains_any', [region]).limit(1).stream()
    wild_type_seq = None
    for obj in wild_type:
        wild_type_seq = obj.to_dict()['seq']
    return wild_type_seq


def create_xls_file(
        database,
        regions_map
):
    workbook = Workbook()
    rCRS_seq_part, RSRS_seq_part = get_base_sequences_parts(database)

    for region, regionCode in regions_map.items():
        region_sequences = get_sequences_by_region(database, region)
        print(region_sequences)
        wild_type = get_wild_type(region_sequences)
        print(wild_type)
        load_content(workbook, region_sequences, rCRS_seq_part, RSRS_seq_part, wild_type, regionCode)

    region_sequences = get_sequences_not_in_regions(database, list(regions_map.keys()))
    print(region_sequences)
    wild_type = get_wild_type(region_sequences)
    load_content(workbook, region_sequences, rCRS_seq_part, RSRS_seq_part, wild_type, "Rest")

    workbook.save(filename=f"mtDNA.xlsx")


def load_content(workbook, region_sequences, rCRS_seq_part, RSRS_seq_part, wild_type, regionCode):
    rcrs_distribution = base_distances(region_sequences, rCRS_seq_part)
    rsrs_distribution = base_distances(region_sequences, RSRS_seq_part)
    wild_type_distribution = base_distances(region_sequences, wild_type)
    pair_distribution = pair_distances(region_sequences)
    print(pair_distribution)

    sheet = workbook.create_sheet(regionCode)
    statistics_template = ['Мат сподів.', 'Сер. квадр. відхил.', 'Мода', 'Мін.', 'Макс.', 'Коеф. варіац.']
    sheet.cell(row=1, column=1, value='Відстань')
    for i in range(max_distance):
        sheet.cell(row=1, column=i + 2, value=i)

    # rCRS
    rCRS_histogram, rCRS_parts = get_distribution_histograms(rcrs_distribution)

    sheet.cell(row=2, column=1, value='Розподіл відносно базової rCRS')
    for i in range(len(rCRS_histogram)):
        sheet.cell(row=2, column=i + 2, value=rCRS_histogram[i])

    sheet.cell(row=3, column=1, value='Розподіл відносно базової rCRS (частка)')
    for i in range(len(rCRS_parts)):
        sheet.cell(row=3, column=i + 2, value=rCRS_parts[i])

    for i in range(len(statistics_template)):
        sheet.cell(row=4, column=i + 2, value=statistics_template[i])

    rCRS_stats = sequence_stats(rcrs_distribution)

    for i in range(len(rCRS_stats)):
        sheet.cell(row=5, column=i + 2, value=rCRS_stats[i])

    # RSRS
    RSRS_histogram, RSRS_parts = get_distribution_histograms(rsrs_distribution)

    sheet.cell(row=6, column=1, value='Розподіл відносно базової RSRS')
    for i in range(len(RSRS_histogram)):
        sheet.cell(row=6, column=i + 2, value=RSRS_histogram[i])

    sheet.cell(row=7, column=1, value='Розподіл відносно базової RSRS (частка)')
    for i in range(len(RSRS_parts)):
        sheet.cell(row=7, column=i + 2, value=RSRS_parts[i])

    for i in range(len(statistics_template)):
        sheet.cell(row=8, column=i + 2, value=statistics_template[i])

    rsrs_stats = sequence_stats(rsrs_distribution)

    for i in range(len(rsrs_stats)):
        sheet.cell(row=9, column=i + 2, value=rsrs_stats[i])

    # Wild type

    wild_type_histogram, wild_type_parts = get_distribution_histograms(wild_type_distribution)

    sheet.cell(row=10, column=1, value='Розподіл відносно дикого типу')
    for i in range(len(wild_type_histogram)):
        sheet.cell(row=10, column=i + 2, value=wild_type_histogram[i])

    sheet.cell(row=11, column=1, value='Розподіл відносно дикого типу (частка)')
    for i in range(len(wild_type_parts)):
        sheet.cell(row=11, column=i + 2, value=wild_type_parts[i])

    for i in range(len(statistics_template)):
        sheet.cell(row=12, column=i + 2, value=statistics_template[i])

    wild_type_stats = sequence_stats(wild_type_distribution)

    for i in range(len(wild_type_stats)):
        sheet.cell(row=13, column=i + 2, value=wild_type_stats[i])

    # Paired

    pair_histogram, pair_parts = get_distribution_histograms(pair_distribution)

    sheet.cell(row=14, column=1, value='Розподіл відносно попарних')
    for i in range(len(pair_histogram)):
        sheet.cell(row=14, column=i + 2, value=pair_histogram[i])

    sheet.cell(row=15, column=1, value='Розподіл відносно попарних (частка)')
    for i in range(len(pair_parts)):
        sheet.cell(row=15, column=i + 2, value=pair_parts[i])

    for i in range(len(statistics_template)):
        sheet.cell(row=16, column=i + 2, value=statistics_template[i])

    pair_stats = sequence_stats(pair_distribution)

    for i in range(len(pair_stats)):
        sheet.cell(row=17, column=i + 2, value=pair_stats[i])

    sheet.cell(row=18, column=2, value='Рядочок дикого типу')
    sheet.cell(row=18, column=3, value=wild_type)
    sheet.cell(row=19, column=2, value='Кількість поліморфізмів у дикого типу відносно базової rCRS')
    wild_type_rCRS_distance = base_distances([wild_type], rCRS_seq_part)[0]
    sheet.cell(row=19, column=3, value=wild_type_rCRS_distance)
    sheet.cell(row=20, column=2, value='Кількість поліморфізмів у дикого типу відносно базової RSRS')
    wild_type_RSRS_distance = base_distances([wild_type], RSRS_seq_part)[0]
    sheet.cell(row=20, column=3, value=wild_type_RSRS_distance)
    sheet.cell(row=22, column=2, value='Кількість поліморфізмів у популяції відносно базової rCRS')
    total_rCRS = sum(rcrs_distribution)
    sheet.cell(row=22, column=3, value=total_rCRS)
    sheet.cell(row=23, column=2, value='Кількість поліморфізмів у популяції відносно базової RSRS')
    total_RSRS = sum(rsrs_distribution)
    sheet.cell(row=23, column=3, value=total_RSRS)


# Coefficient of variation
cv = lambda x: np.std(x, ddof=1) / np.mean(x) * 100


def sequence_stats(arr):
    expected_value = mean(arr)
    standard_error = np.std(arr, ddof=1) / np.sqrt(np.size(arr))
    distance_mode = mode(arr)
    distance_min = min(arr)
    distance_max = max(arr)
    distance_variation = cv(arr)
    return [expected_value, standard_error, distance_mode, distance_min, distance_max, distance_variation]
